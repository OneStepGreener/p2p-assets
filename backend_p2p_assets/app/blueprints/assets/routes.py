import io
import csv
from flask import request, jsonify, Response, send_file
from flask_restx import Resource
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.blueprints.assets import assets_bp, assets_ns
from app.blueprints.assets.services import AssetService
from app.blueprints.assets import queries as asset_queries
from app.extensions import db
from app.core.auth_scope import get_current_scope
from app.utils.helpers import build_response, db_error_message_and_code


@assets_ns.route('/kpis')
@assets_ns.doc('get_kpis')
class AssetKPIs(Resource):
    """Get asset KPIs including total active count and total cost"""
    
    def get(self):
        """Get asset KPIs"""
        try:
            # Get total count from srm_assets table
            total_active_count = AssetService.get_total_active_count()
            # Get total cost sum from srm_assets table
            total_cost_sum = AssetService.get_total_cost_sum()
            
            return build_response(
                data={
                    'total_active_assets': total_active_count,
                    'total_assets': total_active_count,  # All records count as active
                    'total_gross_value': total_cost_sum
                },
                message='KPIs retrieved successfully',
                status_code=200
            ), 200
            
        except Exception as e:
            return build_response(
                message=f'Failed to retrieve KPIs: {str(e)}',
                status_code=500
            ), 500


@assets_ns.route('/transfer')
@assets_ns.doc('transfer_asset')
class AssetTransfer(Resource):
    """
    Initiate asset transfer.

    POST body:
    {
        "assetId": "...",
        "destinationBranchCode": "...",
        "reason": "...",
        "refNo": "..."
    }
    """

    @jwt_required()
    def post(self):
        try:
            body = request.get_json(silent=True) or {}
            asset_id = (body.get('assetId') or '').strip()
            dest_branch = (body.get('destinationBranchCode') or '').strip()
            reason = (body.get('reason') or '').strip()
            remarks = (body.get('remarks') or '').strip()
            ref_no = (body.get('refNo') or '').strip()

            if not asset_id:
                return build_response(message='assetId is required', status_code=400), 400
            if not dest_branch:
                return build_response(message='destinationBranchCode is required', status_code=400), 400
            if not reason:
                return build_response(message='reason is required', status_code=400), 400
            if not ref_no:
                return build_response(message='refNo is required', status_code=400), 400

            current_user = get_jwt_identity() or ''
            # Use the current user as both transferred_by and approved_by by default.
            AssetService.transfer_asset(
                asset_id=asset_id,
                destination_branch_code=dest_branch,
                reason=reason,
                ref_no=ref_no,
                transferred_by=current_user,
                approved_by=current_user,
                transfer_remarks=remarks or reason,
            )

            return build_response(
                data={'assetId': asset_id},
                message='Asset transfer initiated successfully',
                status_code=200,
            ), 200
        except Exception as e:
            db.session.rollback()
            return build_response(
                message=f'Failed to initiate transfer: {str(e)}',
                status_code=500,
            ), 500


@assets_ns.route('/count')
@assets_ns.doc('get_count')
class AssetCount(Resource):
    """Get count of assets; optional query params (divisionCode, branchCode, search, etc.) filter like register."""

    @jwt_required()
    def get(self):
        """Get asset count; with no params returns all; with params returns filtered count."""
        try:
            scope = get_current_scope()
            params = _register_params(scope=scope)
            count = AssetService.get_total_active_count(params=params)
            return build_response(data={'count': count}, message='Count retrieved successfully', status_code=200), 200
        except Exception as e:
            return build_response(message=f'Failed to retrieve count: {str(e)}', status_code=500), 500


@assets_bp.route('/count', methods=['GET'])
@jwt_required()
def get_asset_count():
    """Asset count; query params (divisionCode, branchCode, search, serviceCode, etc.) filter like register."""
    try:
        scope = get_current_scope()
        params = _register_params(scope=scope)
        count = AssetService.get_total_active_count(params=params)
        return jsonify(success=True, data={'count': count}, message='Count retrieved successfully'), 200
    except Exception as e:
        msg, code = db_error_message_and_code(e)
        return jsonify(success=False, message=msg if code == 503 else f'Failed to retrieve count: {str(e)}'), code


@assets_bp.route('/transfer', methods=['POST'])
@jwt_required()
def transfer_asset():
    """
    Initiate asset transfer (blueprint route variant).

    Body:
    {
        "assetId": "...",
        "destinationBranchCode": "...",
        "reason": "...",
        "refNo": "..."
    }
    """
    try:
        body = request.get_json(silent=True) or {}
        asset_id = (body.get('assetId') or '').strip()
        dest_branch = (body.get('destinationBranchCode') or '').strip()
        reason = (body.get('reason') or '').strip()
        remarks = (body.get('remarks') or '').strip()
        ref_no = (body.get('refNo') or '').strip()

        if not asset_id:
            return jsonify(success=False, message='assetId is required'), 400
        if not dest_branch:
            return jsonify(success=False, message='destinationBranchCode is required'), 400
        if not reason:
            return jsonify(success=False, message='reason is required'), 400
        if not ref_no:
            return jsonify(success=False, message='refNo is required'), 400

        current_user = get_jwt_identity() or ''
        AssetService.transfer_asset(
            asset_id=asset_id,
            destination_branch_code=dest_branch,
            reason=reason,
            ref_no=ref_no,
            transferred_by=current_user,
            approved_by=current_user,
            transfer_remarks=remarks or reason,
        )

        return jsonify(
            success=True,
            data={'assetId': asset_id},
            message='Asset transfer initiated successfully',
        ), 200
    except Exception as e:
        db.session.rollback()
        return jsonify(
            success=False,
            message=f'Failed to initiate transfer: {str(e)}',
        ), 500


@assets_ns.route('/total-cost')
@assets_ns.doc('get_total_cost')
class TotalCost(Resource):
    """Get total cost sum; optional query params filter like register."""

    @jwt_required()
    def get(self):
        try:
            scope = get_current_scope()
            params = _register_params(scope=scope)
            total_cost = AssetService.get_total_cost_sum(params=params)
            return build_response(data={'total_cost': total_cost}, message='Total cost retrieved successfully', status_code=200), 200
        except Exception as e:
            return build_response(message=f'Failed to retrieve total cost: {str(e)}', status_code=500), 500


@assets_bp.route('/total-cost', methods=['GET'])
@jwt_required()
def get_total_cost():
    """Total cost; query params filter like register."""
    try:
        scope = get_current_scope()
        params = _register_params(scope=scope)
        total_cost = AssetService.get_total_cost_sum(params=params)
        return jsonify(success=True, data={'total_cost': total_cost}, message='Total cost retrieved successfully'), 200
    except Exception as e:
        msg, code = db_error_message_and_code(e)
        return jsonify(success=False, message=msg if code == 503 else f'Failed to retrieve total cost: {str(e)}'), code


@assets_ns.route('/net-book-value')
@assets_ns.doc('get_net_book_value')
class NetBookValue(Resource):
    """Get net book value; optional query params filter like register."""

    @jwt_required()
    def get(self):
        try:
            scope = get_current_scope()
            params = _register_params(scope=scope)
            net_book_value = AssetService.get_net_book_value(params=params)
            return build_response(data={'net_book_value': net_book_value}, message='Net book value retrieved successfully', status_code=200), 200
        except Exception as e:
            return build_response(message=f'Failed to retrieve net book value: {str(e)}', status_code=500), 500


@assets_bp.route('/net-book-value', methods=['GET'])
@jwt_required()
def get_net_book_value():
    """Net book value; query params filter like register."""
    try:
        scope = get_current_scope()
        params = _register_params(scope=scope)
        net_book_value = AssetService.get_net_book_value(params=params)
        return jsonify(success=True, data={'net_book_value': net_book_value}, message='Net book value retrieved successfully'), 200
    except Exception as e:
        msg, code = db_error_message_and_code(e)
        return jsonify(success=False, message=msg if code == 503 else f'Failed to retrieve net book value: {str(e)}'), code


@assets_ns.route('')
@assets_ns.doc('get_assets')
class AssetsList(Resource):
    """GET /api/v1/assets - Asset Register list with search, pagination, sort. Role-based: USER sees only own scope."""

    @jwt_required()
    def get(self):
        """Get assets (same as /register). Query: search, page, pageSize, sortBy (assetId|assetName|acquisitionDate|cost|status), sortDir (asc|desc)."""
        try:
            scope = get_current_scope()
            params = _register_params(scope=scope)
            result = AssetService.get_asset_register(params)
            resp = build_response(
                data=result['data'],
                message='Assets retrieved successfully',
                status_code=200
            )
            resp['page'] = result['page']
            resp['pageSize'] = result['pageSize']
            resp['totalRows'] = result['totalRows']
            return resp, 200
        except Exception as e:
            return build_response(
                message=f'Failed to retrieve assets: {str(e)}',
                status_code=500
            ), 500


def _register_params(scope=None):
    """Parse query params for Asset Register (search, hierarchy filters, dates, pagination, sort). Adds scope for role-based filter."""
    params = {
        'search': request.args.get('search', '').strip() or None,
        'country': request.args.get('country', '').strip() or None,
        'state': request.args.get('state', '').strip() or None,
        'city': request.args.get('city', '').strip() or None,
        'companyCode': request.args.get('companyCode', '').strip() or None,
        'divCode': request.args.get('divCode', '').strip() or None,
        'divisionCode': request.args.get('divisionCode', '').strip() or None,
        'regionCode': request.args.get('regionCode', '').strip() or None,
        'branchCode': request.args.get('branchCode', '').strip() or None,
        'serviceCode': request.args.get('serviceCode', '').strip() or None,
        'categoryId': request.args.get('categoryId', '').strip() or None,
        'subcategoryId': request.args.get('subcategoryId', '').strip() or None,
        'status': request.args.get('status', '').strip() or None,
        'fromDate': request.args.get('fromDate', '').strip() or None,
        'toDate': request.args.get('toDate', '').strip() or None,
        'page': request.args.get('page', '1'),
        'pageSize': request.args.get('pageSize', '20'),
        'sortBy': request.args.get('sortBy', '').strip() or None,
        'sortDir': request.args.get('sortDir', 'asc').strip() or 'asc',
    }
    if scope:
        params['scope'] = scope
    return params


@assets_ns.route('/mark-idle')
@assets_ns.doc('mark_asset_idle')
class MarkAssetIdle(Resource):
    """Mark an asset as idle; stores ASSET_REF_ID and IDLE_REASON in SRM_IDLE_ASSETS."""

    @jwt_required()
    def post(self):
        """POST body: { "assetId": "...", "reason": "..." }. Inserts into SRM_IDLE_ASSETS."""
        try:
            body = request.get_json(silent=True) or {}
            asset_id = (body.get('assetId') or '').strip()
            reason = (body.get('reason') or '').strip()
            if not asset_id:
                return build_response(message='assetId is required', status_code=400), 400
            if not reason:
                return build_response(message='reason is required', status_code=400), 400
            AssetService.mark_asset_idle(asset_id, reason)
            return build_response(
                data={'assetId': asset_id},
                message='Asset marked as idle successfully',
                status_code=200
            ), 200
        except Exception as e:
            db.session.rollback()
            return build_response(
                message=f'Failed to mark asset as idle: {str(e)}',
                status_code=500
            ), 500


@assets_ns.route('/mark-active')
@assets_ns.doc('mark_asset_active')
class MarkAssetActive(Resource):
    """Mark an asset as active; sets ASSET_STATUS = 'Active' in srm_assets."""

    @jwt_required()
    def post(self):
        """POST body: { "assetId": "..." }. Updates srm_assets.ASSET_STATUS to Active."""
        try:
            body = request.get_json(silent=True) or {}
            asset_id = (body.get('assetId') or '').strip()
            if not asset_id:
                return build_response(message='assetId is required', status_code=400), 400
            AssetService.mark_asset_active(asset_id)
            return build_response(
                data={'assetId': asset_id},
                message='Asset marked as active successfully',
                status_code=200
            ), 200
        except Exception as e:
            db.session.rollback()
            return build_response(
                message=f'Failed to mark asset as active: {str(e)}',
                status_code=500
            ), 500


@assets_ns.route('/register')
@assets_ns.doc('get_asset_register')
class AssetRegister(Resource):
    """Asset Register from srm_assets (single source of truth) with label joins. Role-based: USER sees only own div/branch/region."""

    @jwt_required()
    def get(self):
        """Get paginated Asset Register with search, filters, sort. Requires JWT; USER scope applied in backend."""
        try:
            scope = get_current_scope()
            params = _register_params(scope=scope)
            result = AssetService.get_asset_register(params)
            resp = build_response(
                data=result['data'],
                message='Asset register retrieved successfully',
                status_code=200
            )
            resp['page'] = result['page']
            resp['pageSize'] = result['pageSize']
            resp['totalRows'] = result['totalRows']
            return resp, 200
        except Exception as e:
            return build_response(
                message=f'Failed to retrieve asset register: {str(e)}',
                status_code=500
            ), 500


@assets_bp.route('/mark-idle', methods=['POST'])
@jwt_required()
def mark_asset_idle():
    """Mark an asset as idle; stores in SRM_IDLE_ASSETS. Body: { \"assetId\": \"...\", \"reason\": \"...\" }."""
    try:
        body = request.get_json(silent=True) or {}
        asset_id = (body.get('assetId') or '').strip()
        reason = (body.get('reason') or '').strip()
        if not asset_id:
            return jsonify(success=False, message='assetId is required'), 400
        if not reason:
            return jsonify(success=False, message='reason is required'), 400
        AssetService.mark_asset_idle(asset_id, reason)
        return jsonify(success=True, data={'assetId': asset_id}, message='Asset marked as idle successfully'), 200
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f'Failed to mark asset as idle: {str(e)}'), 500


@assets_bp.route('/mark-active', methods=['POST'])
@jwt_required()
def mark_asset_active():
    """Mark an asset as active. Body: { \"assetId\": \"...\" }."""
    try:
        body = request.get_json(silent=True) or {}
        asset_id = (body.get('assetId') or '').strip()
        if not asset_id:
            return jsonify(success=False, message='assetId is required'), 400
        AssetService.mark_asset_active(asset_id)
        return jsonify(success=True, data={'assetId': asset_id}, message='Asset marked as active successfully'), 200
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=f'Failed to mark asset as active: {str(e)}'), 500


@assets_bp.route('/register', methods=['GET'])
@jwt_required()
def get_asset_register():
    """Asset Register from srm_assets with search, filters, pagination, sort. Role-based filter applied."""
    try:
        scope = get_current_scope()
        params = _register_params(scope=scope)
        result = AssetService.get_asset_register(params)
        return jsonify({
            'success': True,
            'data': result['data'],
            'page': result['page'],
            'pageSize': result['pageSize'],
            'totalRows': result['totalRows'],
            'message': 'Asset register retrieved successfully'
        }), 200
    except Exception as e:
        msg, code = db_error_message_and_code(e)
        return jsonify({
            'success': False,
            'message': msg if code == 503 else f'Failed to retrieve asset register: {msg}'
        }), code


@assets_ns.route('/export')
@assets_ns.doc('export_assets')
class AssetExport(Resource):
    """Export Asset Register as CSV or XLSX (same filters as /register). Role-based filter applied."""

    @jwt_required()
    def get(self):
        """Export filtered assets as CSV or XLSX. Requires JWT; USER scope applied."""
        try:
            fmt = (request.args.get('format') or 'csv').strip().lower()
            if fmt not in ('csv', 'xlsx'):
                return build_response(message='format must be csv or xlsx', status_code=400), 400
            scope = get_current_scope()
            params = _register_params(scope=scope)
            if fmt == 'csv':
                return _stream_csv(params)
            return _send_xlsx(params)
        except Exception as e:
            return build_response(message=f'Export failed: {str(e)}', status_code=500), 500


def _stream_csv(params):
    """Stream CSV response. Columns match Asset Register spec."""
    def generate():
        buf = io.StringIO()
        w = csv.writer(buf)
        headers = [
            'assetId', 'assetName', 'divisionName', 'branchName', 'city', 'state', 'country', 'regionCode',
            'status', 'vendor', 'acquisitionDate', 'putToUseDate', 'poNo', 'invNo',
            'cost', 'totalCost', 'accumDep', 'netBookValue'
        ]
        w.writerow(headers)
        yield buf.getvalue()
        buf.seek(0)
        buf.truncate(0)
        for row in asset_queries.get_asset_register_export(db.session, params):
            w.writerow([
                row.get('assetId', ''),
                row.get('assetName', ''),
                row.get('divisionName', ''),
                row.get('branchName', ''),
                row.get('city', ''),
                row.get('state', ''),
                row.get('country', ''),
                row.get('regionCode', ''),
                row.get('status', ''),
                row.get('vendor', ''),
                row.get('acquisitionDate', ''),
                row.get('putToUseDate', ''),
                row.get('poNo', ''),
                row.get('invNo', ''),
                row.get('cost', 0),
                row.get('totalCost', 0),
                row.get('accumDep', 0),
                row.get('netBookValue', 0),
            ])
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)
    return Response(
        generate(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=assets_register.csv'}
    )


def _send_xlsx(params):
    """Build XLSX in memory and return as attachment. Currency columns formatted."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({'success': False, 'message': 'openpyxl not installed'}), 500
    wb = Workbook()
    ws = wb.active
    ws.title = 'Asset Register'
    headers = [
        'Asset ID', 'Asset Name', 'Division', 'Branch', 'Region', 'City', 'State', 'Country',
        'Status', 'Vendor', 'Acquisition Date', 'Put to Use Date', 'PO No', 'Invoice No',
        'Cost', 'Total Cost', 'Accum Dep', 'Net Book Value'
    ]
    ws.append(headers)
    keys = [
        'assetId', 'assetName', 'divisionName', 'branchName', 'regionCode', 'city', 'state', 'country',
        'status', 'vendor', 'acquisitionDate', 'putToUseDate', 'poNo', 'invNo',
        'cost', 'totalCost', 'accumDep', 'netBookValue'
    ]
    num_cols = (15, 16, 17, 18)  # cost, totalCost, accumDep, netBookValue (1-based)
    for row in asset_queries.get_asset_register_export(db.session, params):
        rw = []
        for i, k in enumerate(keys):
            v = row.get(k, 0 if k in ('cost', 'totalCost', 'accumDep', 'netBookValue') else '')
            rw.append(v)
        ws.append(rw)
    for r in range(2, ws.max_row + 1):
        for c in num_cols:
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).number_format = '#,##0.00'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='assets_register.xlsx'
    )


@assets_bp.route('/export', methods=['GET'])
@jwt_required()
def export_assets():
    """Export Asset Register as CSV or XLSX. Role-based filter applied."""
    try:
        fmt = (request.args.get('format') or 'csv').strip().lower()
        if fmt not in ('csv', 'xlsx'):
            return jsonify({'success': False, 'message': 'format must be csv or xlsx'}), 400
        scope = get_current_scope()
        params = _register_params(scope=scope)
        if fmt == 'csv':
            return _stream_csv(params)
        return _send_xlsx(params)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Export failed: {str(e)}'}), 500


@assets_ns.route('/locations')
@assets_ns.doc('get_locations')
class AssetLocations(Resource):
    """Get distinct HO_COUNTRY_NAME from cm_division_tlog for location filter"""

    def get(self):
        """Get list of locations (distinct HO_COUNTRY_NAME)"""
        try:
            locations = AssetService.get_locations()
            return build_response(
                data=locations,
                message='Locations retrieved successfully',
                status_code=200
            ), 200
        except Exception as e:
            return build_response(
                message=f'Failed to retrieve locations: {str(e)}',
                status_code=500
            ), 500


@assets_ns.route('/category-counts')
@assets_ns.doc('get_category_counts')
class CategoryCounts(Resource):
    """Get category counts by service_code (distinct subcategory count per category)"""

    def get(self):
        try:
            service_code = request.args.get('service_code', '').strip()
            if not service_code:
                return build_response(
                    message='service_code is required',
                    status_code=400
                ), 400

            data = AssetService.get_category_counts_by_service(service_code)
            return build_response(
                data=data,
                message='Category counts retrieved successfully',
                status_code=200
            ), 200
        except Exception as e:
            return build_response(
                message=f'Failed to retrieve category counts: {str(e)}',
                status_code=500
            ), 500


@assets_bp.route('/locations', methods=['GET'])
def get_locations():
    """Get distinct HO_COUNTRY_NAME from cm_division_tlog for location filter"""
    try:
        locations = AssetService.get_locations()
        return jsonify({
            'success': True,
            'data': locations,
            'message': 'Locations retrieved successfully'
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to retrieve locations: {str(e)}'
        }), 500


# GET /api/v1/assets is handled by AssetsList Resource (namespace) with JWT + register format.
# Removed duplicate blueprint route that returned legacy { data: { assets, pagination } } and caused blank page.
